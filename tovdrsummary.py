import os
from openpyxl import load_workbook

# Main directory
main_directory = r'C:\Users\user\PycharmProjects\toVDRSummary\VDR SUMMARY'

# List to store subfolders
subfolders_list = {}

# Iterate through the main directories
for client_folder in ['ENQUEST', 'JADESTONEENERGY', 'PFLNG']:
    # Get the path of the main folder
    main_folder_path = os.path.join(main_directory, client_folder)

    # Check if the main folder exists
    if os.path.exists(main_folder_path):
        # Get list of subfolders in the main folder
        subfolders = [folder for folder in os.listdir(main_folder_path) if
                      os.path.isdir(os.path.join(main_folder_path, folder))]
        # Add the list of subfolders to the dictionary
        subfolders_list[client_folder] = subfolders
    else:
        print("Main folder does not exist:", main_folder_path)

# Process files in subfolders within each main directory
for main_folder, subfolders in subfolders_list.items():
    for subfolder in subfolders:
        # Get the path of the subfolder
        vessel_dir = os.path.join(main_directory, main_folder, subfolder)
        print(f"Processing files in {vessel_dir}:")

        # Process CSV files and change extension to XLSX
        for file in os.listdir(vessel_dir):
            try:
                if file.endswith('.csv'):
                    csv_path = os.path.join(vessel_dir, file)
                    new_xlsx_path = os.path.join(vessel_dir, os.path.splitext(file)[0] + '.xlsx')
                    os.rename(csv_path, new_xlsx_path)
                    print(f"Changed extension of {file} to .xlsx")
                else:
                    print("No .csv file to process:")
            except Exception as e:
                print(f"Error processing file: {file}. Error: {e}")

        # Process Excel files starting with 'Daily Summary Report'
        summary_files = [file for file in os.listdir(vessel_dir) if
                         file.startswith('Daily Summary Report') and file.endswith('.xlsx')]
        destination_file = os.path.join(vessel_dir, 'VDR SUMMARY.xlsx')
        try:
            dest_wb = load_workbook(destination_file)
            summary_ws = dest_wb['Summary']
        except FileNotFoundError:
            print(f"File not found: {destination_file}. Skipping processing for {vessel_dir}")
            continue

        # Processing 'Daily Summary Report' files
        summary_start_row = 2
        summary_start_column = 26

        for file_name in summary_files:
            try:
                # Load Excel file
                excel_path = os.path.join(vessel_dir, file_name)
                excel_wb = load_workbook(excel_path)
                excel_ws = excel_wb.active

                # Get data from Excel file
                data_C2 = excel_ws['C2'].value
                data_H2_J2 = [cell.value for cell in excel_ws['H2:J2'][0]]  # Assuming H2:J2 are merged

                # Paste data into destination workbook
                summary_ws.cell(row=summary_start_row, column=summary_start_row, value=data_C2)
                for idx, cell_value in enumerate(data_H2_J2):
                    summary_ws.cell(row=summary_start_row, column=summary_start_row + idx + 1, value=cell_value)

                # Move to the next row in the destination workbook
                summary_start_row += 1
                print(f"Data from {file_name} copied to SUMMARY sheet in VDR SUMMARY.xlsx")
            except Exception as e:
                print(f"Error processing file: {file_name}. Error: {e}")

        # Process Excel files starting with 'Vessel Working Report'
        trails_files = [file for file in os.listdir(vessel_dir) if
                        file.startswith('Vessel Working Report') and file.endswith('.xlsx')]

        # Processing 'Vessel Working Report' files
        for file in trails_files:
            try:
                # Load the Excel file
                excel_path = os.path.join(vessel_dir, file)
                excel_wb = load_workbook(excel_path)

                # Get the active worksheet
                source_ws = excel_wb.active

                # Load the destination workbook
                dest_ws = dest_wb['Trails']  # Assuming the destination sheet is named 'Trails'

                # Copy data from the active worksheet to Trails sheet
                for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, max_row=500, min_col=1, max_col=6),
                                              start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        dest_ws.cell(row=row_idx, column=col_idx, value=cell.value)

                print(f"Data from {file} copied to TRAILS sheet in VDR SUMMARY.xlsx")
            except Exception as e:
                print(f"Error processing file: {file}. Error: {e}")

        # Save changes in destination workbook
        try:
            dest_wb.save(destination_file)
            print("FINISHED!!.")
        except Exception as e:
            print(f"Error saving changes to destination workbook. Error: {e}")
